import logging
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXCEL_PATH = Path(os.getenv("DATA_DIR", str(Path(__file__).parent.parent / "data"))) / "Informe_Amenazas.xlsx"

THREATS_COLS = [
    "ID", "Fecha Detección", "Fuente", "V-A", "CVE(s) Identificados", "NOMBRE",
    "Descripción Técnica", "Descripción del Riesgo", "TTPs (MITRE ATT&CK)",
    "Probabilidad", "Impacto", "Critcidad", "Acción Recomendada",
    "Área Responsable Remediación", "Fecha Escalamiento al Área",
    "¿Afecta Activos? (Tenable)", "Comentarios Tenable",
    "Bloqueo Antivirus", "Bloqueo Firewall", "Caso Firewall",
]

IOCS_COLS = [
    "ID Amenaza", "Tipo de IoC", "Indicador (IoC)", "Bloqueo Antivirus", "Bloqueo Firewall",
]

_HEADER_FILL       = PatternFill("solid", fgColor="1F3864")
_HEADER_FILL_GREEN = PatternFill("solid", fgColor="375623")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_ROW_FONT     = Font(size=9)
_ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
_BORDER = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)

_SOURCE_PREFIX = {
    "colcert":             "COL",
    "boletin csirt cywex": "WX",
    "csirt cywex":         "WX",
    "wexler":              "WX",
    "octapus":             "OCT",
    "octopus":             "OCT",
}

# Fuentes que son boletines WEXLER/CyWex y para las que se usa el ID capturado del boletín
_WEXLER_SOURCES = {"boletin csirt cywex", "csirt cywex", "wexler"}

# Mapping from old column names to new canonical names
COLUMN_ALIASES = {
    'Tipo de Amenaza':                           'V-A',
    'Fuente de Detección':                       'Fuente',
    'Vulnerabilidad / Amenaza':                  'NOMBRE',
    'Descripción':                               'Descripción Técnica',
    'CRITICIDAD':                                'Critcidad',
    'PROBABILIDAD':                              'Probabilidad',
    'Posible Impacto':                           'Impacto',
    'TTP (Técnicas, Tácticas y Procedimientos)': 'TTPs (MITRE ATT&CK)',
    'Comentarios SOC':                           'Comentarios Tenable',
    'Comentarios FIDU':                          'Comentarios Tenable',
    'FECHA DE ESCALAMIENTO':                     'Fecha Escalamiento al Área',
    'AREA RESPONSABLE DE REMEDIACION ':          'Área Responsable Remediación',
    'AREA RESPONSABLE DE REMEDIACION':           'Área Responsable Remediación',
    'BLOQUEO ANTIVIRUS':                         'Bloqueo Antivirus',
    'BLOQUEO FIREWALL':                          'Bloqueo Firewall',
    'CASO FIREWALL':                             'Caso Firewall',
}


def _ensure_excel():
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Registro de Amenazas"
        ws1.append(THREATS_COLS)

        ws2 = wb.create_sheet("Detalle de IoCs")
        _build_ioc_header(ws2)

        wb.save(EXCEL_PATH)
        logger.info(f"Created new Excel at {EXCEL_PATH}")
    else:
        _migrate_columns()


def _build_ioc_header(ws):
    """Write the two-row merged header for the IoC sheet."""
    # Row 1: two merged sections
    ws.merge_cells("A1:C1")
    ws["A1"] = "DETALLE DE INDICADORES DE COMPROMISO (IoC) — ISO 27001 A.5.7"
    ws.merge_cells("D1:E1")
    ws["D1"] = "ACCIONES DE BLOQUEO"

    for col, fill in [(1, _HEADER_FILL), (4, _HEADER_FILL_GREEN)]:
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[1].height = 28

    # Row 2: column labels
    for col_idx, val in enumerate(IOCS_COLS, 1):
        cell = ws.cell(2, col_idx)
        cell.value = val
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[2].height = 22

    col_widths = {get_column_letter(i + 1): w for i, w in enumerate([14, 18, 44, 18, 18])}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A3"


def _style_data_rows(ws, start_row: int = 2):
    """Apply row font, alignment, border and alternating fill to all data rows."""
    for row_idx in range(start_row, ws.max_row + 1):
        is_alt = (row_idx % 2 == 1)
        for cell in ws[row_idx]:
            cell.font      = _ROW_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = _BORDER
            if is_alt:
                cell.fill = _ALT_FILL


def _migrate_columns():
    """Ensure main sheet has the new column set.
    When old column names are detected, performs a full data migration:
    renames columns, remaps data, and rewrites the sheet in canonical order.
    """
    wb = load_workbook(EXCEL_PATH)

    if "Registro de Amenazas" not in wb.sheetnames:
        ws1 = wb.create_sheet("Registro de Amenazas", 0)
        ws1.append(THREATS_COLS)
        _style_sheet_threats(ws1)
        logger.info("Migración: hoja 'Registro de Amenazas' creada")
        wb.save(EXCEL_PATH)
        return

    ws1 = wb["Registro de Amenazas"]
    headers = [ws1.cell(1, c).value for c in range(1, ws1.max_column + 1)]
    changed = False

    # ── Full data migration when old column names are detected ─────────────────
    old_cols_present = [h for h in headers if h and h in COLUMN_ALIASES]
    if old_cols_present:
        logger.info(f"Migración: columnas antiguas detectadas: {old_cols_present}")

        # Read all data rows into dicts, applying alias remapping
        data_rows: list[dict] = []
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and v != "" for v in row):
                continue
            row_dict: dict = {}
            for i, val in enumerate(row):
                if i >= len(headers) or headers[i] is None:
                    continue
                col = str(headers[i]).strip()
                new_col = COLUMN_ALIASES.get(col, col)
                # Columns mapped to None are intentionally dropped
                if new_col is None:
                    continue
                # Prefer non-empty values when two columns map to the same target
                existing = row_dict.get(new_col)
                has_val = val is not None and val != ""
                existing_empty = existing is None or existing == ""
                if new_col not in row_dict or (has_val and existing_empty):
                    row_dict[new_col] = val
            data_rows.append(row_dict)

        # Rewrite the sheet from scratch
        ws1.delete_rows(1, ws1.max_row)
        ws1.append(THREATS_COLS)
        for row_dict in data_rows:
            ws1.append([row_dict.get(c) or "" for c in THREATS_COLS])

        _style_sheet_threats(ws1)
        _style_data_rows(ws1)
        logger.info(f"Migración completa: {len(data_rows)} filas migradas al nuevo formato")
        changed = True

    else:
        # Just add any missing canonical columns at the end
        for col_name in THREATS_COLS:
            if col_name not in headers:
                ws1.cell(1, ws1.max_column + 1).value = col_name
                headers.append(col_name)
                logger.info(f"Migración: columna '{col_name}' agregada")
                changed = True

    # ── Ensure IoC sheet has double header ─────────────────────────────────────
    if "Detalle de IoCs" not in wb.sheetnames:
        ws2 = wb.create_sheet("Detalle de IoCs")
        _build_ioc_header(ws2)
        logger.info("Migración: hoja 'Detalle de IoCs' creada")
        changed = True
    else:
        ws2 = wb["Detalle de IoCs"]
        if ws2.cell(1, 1).value != "DETALLE DE INDICADORES DE COMPROMISO (IoC) — ISO 27001 A.5.7":
            if ws2.max_row >= 1:
                ws2.insert_rows(1)
            _build_ioc_header(ws2)
            logger.info("Migración: encabezado doble de IoCs aplicado")
            changed = True

    if changed:
        wb.save(EXCEL_PATH)


def _style_sheet_threats(ws):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 28

    widths = [10, 14, 22, 18, 22, 24, 40, 32, 30, 14, 28, 14, 30, 26, 20, 22, 36, 18, 18, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A2"


def _next_id(source: str = "", numero_boletin: str = "") -> str:
    """
    Genera el ID para el registro de amenaza.

    Si la fuente es WEXLER/CyWex y se capturó el número de boletín,
    devuelve WX-<numero_boletin> (ej: WX-251).
    En caso contrario genera el siguiente ID secuencial.
    """
    src = source.lower().strip()
    prefix = "WX"
    for key, pre in _SOURCE_PREFIX.items():
        if key in src:
            prefix = pre
            break

    # Usar el número de boletín capturado para fuentes WEXLER
    if numero_boletin and any(key in src for key in _WEXLER_SOURCES):
        candidate = f"WX-{numero_boletin}"
        # Si ya existe ese ID, añadir sufijo para evitar duplicado
        try:
            df = pd.read_excel(EXCEL_PATH, sheet_name="Registro de Amenazas", engine="openpyxl")
            existing = set(df["ID"].dropna().astype(str).str.strip()) if "ID" in df.columns else set()
            if candidate not in existing:
                return candidate
            # Sufijo incremental si el ID ya está registrado
            suffix = 2
            while f"{candidate}-{suffix}" in existing:
                suffix += 1
            return f"{candidate}-{suffix}"
        except Exception:
            return candidate

    pattern = re.compile(rf"{re.escape(prefix)}-(\d+)$")
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Registro de Amenazas", engine="openpyxl")
        if df.empty or "ID" not in df.columns:
            return f"{prefix}-001"
        nums = []
        for raw in df["ID"].dropna().astype(str):
            m = pattern.match(raw.strip())
            if m:
                n = int(m.group(1))
                if n < 90000:
                    nums.append(n)
        if not nums:
            return f"{prefix}-001"
        return f"{prefix}-{max(nums) + 1:03d}"
    except Exception as exc:
        logger.warning(f"_next_id error: {exc}")
        return f"{prefix}-001"


def add_record(analysis: dict, pdf_name: str) -> str:
    _ensure_excel()
    source          = analysis.get("fuente", "")
    numero_boletin  = str(analysis.get("numero_boletin", "")).strip()
    threat_id       = _next_id(source, numero_boletin)
    today = datetime.now().strftime("%Y-%m-%d")

    tipo = analysis.get("tipo_amenaza", "Otro")
    nombre = analysis.get("nombre", "") if tipo == "Vulnerabilidad" else "NO APLICA"

    threat_row = {
        "ID":                        threat_id,
        "Fecha Detección":           today,
        "Fuente":                    analysis.get("fuente", "CARONTE"),
        "V-A":                       tipo,
        "CVE(s) Identificados":      analysis.get("cves", ""),
        "NOMBRE":                    nombre,
        "Descripción Técnica":       analysis.get("descripcion", ""),
        "Descripción del Riesgo":    analysis.get("riesgo", ""),
        "TTPs (MITRE ATT&CK)":       analysis.get("ttps", ""),
        "Probabilidad":              analysis.get("probabilidad", ""),
        "Impacto":                   analysis.get("impacto", ""),
        "Critcidad":                 analysis.get("criticidad", ""),
        "Acción Recomendada":        analysis.get("accion", ""),
        "Área Responsable Remediación": analysis.get("area_responsable_remediacion", ""),
        "Fecha Escalamiento al Área": analysis.get("fecha_escalamiento", ""),
        "¿Afecta Activos? (Tenable)": analysis.get("_tenable_afecta", ""),
        "Comentarios Tenable":       analysis.get("_tenable_comment", ""),
        "Bloqueo Antivirus":         "",
        "Bloqueo Firewall":          "",
        "Caso Firewall":             "",
    }

    iocs_detalle = analysis.get("iocs_detalle", [])
    ioc_rows = [
        {
            "ID Amenaza":       threat_id,
            "Tipo de IoC":      ioc.get("tipo", ""),
            "Indicador (IoC)":  ioc.get("valor", ""),
            "Bloqueo Antivirus": "",
            "Bloqueo Firewall":  "",
        }
        for ioc in iocs_detalle if ioc.get("valor")
    ]

    wb = load_workbook(EXCEL_PATH)

    # ── Threats sheet ──────────────────────────────────────────────────────────
    ws1 = wb["Registro de Amenazas"]
    actual_cols = [ws1.cell(1, c).value for c in range(1, ws1.max_column + 1)]
    ws1.append([threat_row.get(c, "") for c in actual_cols])
    new_row = ws1.max_row
    is_alt  = (new_row % 2 == 1)
    for cell in ws1[new_row]:
        cell.font      = _ROW_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = _BORDER
        if is_alt:
            cell.fill = _ALT_FILL

    _style_sheet_threats(ws1)

    # ── IoCs sheet ─────────────────────────────────────────────────────────────
    ws2 = wb["Detalle de IoCs"]
    for row in ioc_rows:
        ws2.append([row.get(c, "") for c in IOCS_COLS])
        for cell in ws2[ws2.max_row]:
            cell.font      = _ROW_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = _BORDER

    wb.save(EXCEL_PATH)
    logger.info(f"add_record: saved ID={threat_id} ({pdf_name})")
    return threat_id
