"""
Script de migración: YAML + Excel → Oracle 21c

Ejecutar UNA sola vez después de configurar las variables de entorno Oracle:
    python db/migrate_to_oracle.py

Requisitos:
    - Base de datos Oracle 21c accesible y schema creado (db/schema.sql ejecutado).
    - Variables de entorno: ORACLE_USER, ORACLE_PASSWORD, ORACLE_HOST, ORACLE_PORT, ORACLE_SERVICE
    - Archivos fuente: users.yaml y data/Informe_Amenazas.xlsx
"""
import sys
import os
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

import bcrypt
import yaml
import pandas as pd
import math


def migrate():
    from db.oracle_manager import OracleManager

    print("=== CARONTE — Migración a Oracle 21c ===\n")
    db = OracleManager()

    # ── 1. Migrar usuarios (users.yaml) ──────────────────────────────────────
    users_file = ROOT / "users.yaml"
    if users_file.exists():
        with open(users_file, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        users = data.get("users", {})
        migrated_users = 0
        for idx, (username, info) in enumerate(users.items()):
            if db.user_exists(username):
                print(f"  [skip] Usuario '{username}' ya existe en Oracle.")
                continue
            role = "admin" if idx == 0 else "user"
            db.create_user(username, info["password_hash"], role=role)
            if info.get("totp_secret"):
                db.update_totp_secret(username, info["totp_secret"])
            print(f"  [OK]   Usuario '{username}' migrado (role={role})")
            migrated_users += 1
        print(f"\n  Usuarios migrados: {migrated_users}/{len(users)}")
    else:
        print("  [WARN] users.yaml no encontrado, omitiendo migración de usuarios.")

    # ── 2. Migrar amenazas (Informe_Amenazas.xlsx) ────────────────────────────
    excel_path = ROOT / "data" / "Informe_Amenazas.xlsx"
    if excel_path.exists():
        print("\nMigrando amenazas desde Excel…")
        try:
            df_threats = pd.read_excel(excel_path, sheet_name="Registro de Amenazas", engine="openpyxl")
            df_threats = df_threats.where(pd.notnull(df_threats), None)
            migrated_threats = 0
            for _, row in df_threats.iterrows():
                record = {k: (None if isinstance(v, float) and math.isnan(v) else v)
                          for k, v in row.to_dict().items()}
                try:
                    db.upsert_threat(record)
                    migrated_threats += 1
                except Exception as exc:
                    print(f"  [ERR]  Amenaza ID={record.get('ID', '?')}: {exc}")
            print(f"  Amenazas migradas: {migrated_threats}/{len(df_threats)}")
        except Exception as exc:
            print(f"  [ERR]  No se pudo leer la hoja de amenazas: {exc}")

        print("\nMigrando IoCs desde Excel…")
        try:
            # Detectar si tiene doble encabezado
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            ws = wb["Detalle de IoCs"]
            header_row = 1 if "DETALLE" in str(ws.cell(1, 1).value or "").upper() else 0
            wb.close()

            df_iocs = pd.read_excel(excel_path, sheet_name="Detalle de IoCs",
                                    engine="openpyxl", header=header_row)
            df_iocs = df_iocs.where(pd.notnull(df_iocs), None)
            migrated_iocs = 0
            for _, row in df_iocs.iterrows():
                record = {k: (None if isinstance(v, float) and math.isnan(v) else v)
                          for k, v in row.to_dict().items()}
                try:
                    db.add_ioc(record)
                    migrated_iocs += 1
                except Exception as exc:
                    print(f"  [ERR]  IoC: {exc}")
            print(f"  IoCs migrados: {migrated_iocs}/{len(df_iocs)}")
        except Exception as exc:
            print(f"  [ERR]  No se pudo leer la hoja de IoCs: {exc}")
    else:
        print("  [WARN] Informe_Amenazas.xlsx no encontrado, omitiendo migración de datos.")

    print("\n=== Migración completada ===")


if __name__ == "__main__":
    migrate()
