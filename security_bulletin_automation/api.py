import base64
import json
import os
import subprocess
import sys
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, request, send_file, session
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename

load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))
from auth_manager import AuthManager

_IS_PROD = os.getenv("FLASK_ENV", "development") == "production"

app = Flask(__name__)

_secret_key = os.getenv("SECRET_KEY")
if not _secret_key:
    raise RuntimeError("SECRET_KEY no está definida en las variables de entorno. Agrégala al .env.")
app.secret_key = _secret_key
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = _IS_PROD
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=20)

_cors_origins = [o.strip() for o in os.getenv("CORS_ORIGINS", "http://localhost:5173").split(",") if o.strip()]
CORS(app, supports_credentials=True, origins=_cors_origins)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://",
)

_default_data = Path(__file__).parent / "data"
BULLETINS_DIR = Path(os.getenv("DATA_DIR", str(_default_data)))
EXCEL_PATH    = BULLETINS_DIR / "Informe_Amenazas.xlsx"
MAIN_SCRIPT   = Path(__file__).parent / "main.py"

auth = AuthManager()

_USE_ORACLE  = os.getenv("USE_ORACLE",  "false").lower() == "true"
_DISABLE_2FA = os.getenv("DISABLE_2FA", "false").lower() == "true"

_SESSION_TIMEOUT = 20 * 60  # segundos


@app.before_request
def _refresh_session():
    """Cierra la sesión tras 20 minutos de inactividad (sliding timeout)."""
    if not session.get("authenticated"):
        return
    now = datetime.utcnow().timestamp()
    last = session.get("_last_activity", now)
    if now - last > _SESSION_TIMEOUT:
        session.clear()
        from flask import request as _req
        if _req.path.startswith("/api/"):
            from flask import jsonify as _json
            return _json({"error": "Sesión expirada por inactividad"}), 401
    session["_last_activity"] = now


@app.after_request
def _security_headers(response):
    """Elimina información del servidor y previene caché de respuestas de API."""
    response.headers["Server"] = ""
    response.headers["X-Powered-By"] = ""
    if request.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    return response


def _get_oracle():
    from db.oracle_manager import OracleManager
    return OracleManager()


# ── Auth guards ───────────────────────────────────────────────────────────────
def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("authenticated"):
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return wrapper


def require_admin(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("authenticated"):
            return jsonify({"error": "Unauthorized"}), 401
        if session.get("role") != "admin":
            return jsonify({"error": "Se requieren permisos de administrador"}), 403
        return f(*args, **kwargs)
    return wrapper


# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
@limiter.limit("10 per minute; 3 per second")
def login():
    data     = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not auth.has_users():
        auth.create_user(username, password)
        if _DISABLE_2FA:
            session.permanent        = True
            session["authenticated"] = True
            session["username"]      = username
            session["role"]          = auth.get_user_role(username)
            return jsonify({"step": "done", "username": username, "role": session["role"]})
        session["temp_username"] = username
        secret = auth.generate_totp_secret(username)
        qr     = auth.get_qr_image(username, secret)
        return jsonify({
            "step":   "totp_setup",
            "qr":     base64.b64encode(qr).decode(),
            "secret": secret,
        })

    if not auth.verify_password(username, password):
        return jsonify({"error": "Credenciales incorrectas"}), 401

    if _DISABLE_2FA:
        session.permanent        = True
        session["authenticated"] = True
        session["username"]      = username
        session["role"]          = auth.get_user_role(username)
        return jsonify({"step": "done", "username": username, "role": session["role"]})

    session["temp_username"] = username

    if not auth.has_totp(username):
        secret = auth.generate_totp_secret(username)
        qr     = auth.get_qr_image(username, secret)
        return jsonify({
            "step":   "totp_setup",
            "qr":     base64.b64encode(qr).decode(),
            "secret": secret,
        })

    return jsonify({"step": "totp"})


@app.route("/api/auth/totp/verify", methods=["POST"])
def verify_totp():
    data     = request.get_json()
    username = session.get("temp_username")
    code     = data.get("code", "").strip()

    if not username:
        return jsonify({"error": "No hay sesión pendiente"}), 400

    if auth.verify_totp(username, code):
        session.pop("temp_username", None)
        session.permanent        = True
        session["authenticated"] = True
        session["username"]      = username
        session["role"]          = auth.get_user_role(username)
        return jsonify({"step": "done", "username": username, "role": session["role"]})

    return jsonify({"error": "Código incorrecto"}), 401


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/me")
def me():
    if session.get("authenticated"):
        return jsonify({
            "authenticated": True,
            "username": session["username"],
            "role": session.get("role", "user"),
        })
    return jsonify({"authenticated": False})


# ── Admin: gestión de usuarios ────────────────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@require_admin
def list_users():
    return jsonify(auth.list_users())


@app.route("/api/admin/users", methods=["POST"])
@require_admin
@limiter.limit("5 per minute")
def create_user():
    data     = request.get_json()
    username = (data.get("username") or "").strip()
    password = data.get("password", "")
    role     = data.get("role", "user")

    if not username or not password:
        return jsonify({"error": "username y password son requeridos"}), 400
    if role not in ("admin", "user"):
        return jsonify({"error": "role debe ser 'admin' o 'user'"}), 400
    if len(password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres"}), 400
    if auth.user_exists(username):
        return jsonify({"error": f"El usuario '{username}' ya existe"}), 409

    auth.create_user(username, password, role=role)
    return jsonify({"ok": True, "username": username, "role": role}), 201


# ── Files ─────────────────────────────────────────────────────────────────────
@app.route("/api/files", methods=["GET"])
@require_auth
def list_files():
    BULLETINS_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(BULLETINS_DIR.glob("*.pdf"))
    return jsonify([{"name": f.name, "size": f.stat().st_size} for f in files])


@app.route("/api/files", methods=["POST"])
@require_auth
def upload_files():
    BULLETINS_DIR.mkdir(parents=True, exist_ok=True)
    uploaded = []
    for file in request.files.getlist("files"):
        safe_name = secure_filename(file.filename or "upload.pdf")
        if not safe_name.lower().endswith(".pdf"):
            continue
        dest = BULLETINS_DIR / safe_name
        if dest.exists():
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = BULLETINS_DIR / f"{ts}_{safe_name}"
        file.save(dest)
        uploaded.append(dest.name)
    return jsonify({"uploaded": uploaded})


@app.route("/api/files/<name>", methods=["DELETE"])
@require_auth
def delete_file(name):
    path = BULLETINS_DIR / name
    if path.exists() and path.suffix.lower() == ".pdf":
        path.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404


# ── Process (SSE streaming) ───────────────────────────────────────────────────
@app.route("/api/process", methods=["GET", "POST"])
@require_auth
def process():
    def generate():
        try:
            proc = subprocess.Popen(
                [sys.executable, "-u", str(MAIN_SCRIPT)],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                cwd=str(MAIN_SCRIPT.parent),
                env={**os.environ, "PYTHONUNBUFFERED": "1"},
            )
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    yield f"data: {json.dumps({'line': line})}\n\n"
            proc.wait()
            yield f"data: {json.dumps({'done': True, 'code': proc.returncode})}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Archive ───────────────────────────────────────────────────────────────────
def _read_sheet(sheet_name: str, header: int = 0):
    import math
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, engine="openpyxl", header=header)
    records = df.where(pd.notnull(df), None).to_dict(orient="records")
    return [
        {k: (None if isinstance(v, float) and math.isnan(v) else v) for k, v in row.items()}
        for row in records
    ]

def _iocs_header() -> int:
    """Detecta si la hoja de IoCs usa doble encabezado (row 1 fusionado) o encabezado simple."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH, read_only=True)
        ws = wb["Detalle de IoCs"]
        val = ws.cell(1, 1).value or ""
        wb.close()
        return 1 if "DETALLE" in str(val).upper() else 0
    except Exception:
        return 0


_THREATS_COLS = [
    "ID", "Fecha Detección", "Fuente", "V-A", "CVE(s) Identificados", "NOMBRE",
    "Descripción Técnica", "Descripción del Riesgo", "TTPs (MITRE ATT&CK)",
    "Probabilidad", "Impacto", "Critcidad", "Acción Recomendada",
    "Área Responsable Remediación", "Fecha Escalamiento al Área",
    "¿Afecta Activos? (Tenable)", "Comentarios Tenable",
    "Bloqueo Antivirus", "Bloqueo Firewall", "Caso Firewall",
]
_IOCS_COLS = ["ID Amenaza", "Tipo de IoC", "Indicador (IoC)", "Bloqueo Antivirus", "Bloqueo Firewall"]

def _save_excel_atomic(threats: list, iocs: list):
    """Escribe ambas hojas en un archivo temporal y luego lo renombra (escritura atómica)."""
    import math, tempfile, shutil
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ROW_FONT = Font(size=9)
    ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
    BORDER   = Border(
        left=Side(style="thin",   color="B8CCE4"),
        right=Side(style="thin",  color="B8CCE4"),
        top=Side(style="thin",    color="B8CCE4"),
        bottom=Side(style="thin", color="B8CCE4"),
    )
    HDR_FILL_GREEN = PatternFill("solid", fgColor="375623")
    T_WIDTHS = [10, 14, 22, 18, 22, 24, 40, 32, 30, 14, 28, 14, 30, 26, 20, 22, 36, 18, 18, 16]
    I_WIDTHS = [14, 18, 44, 18, 18]

    def _clean(v):
        return None if isinstance(v, float) and math.isnan(v) else v

    def _style_header(ws, widths):
        ws.row_dimensions[1].height = 30
        for cell in ws[1]:
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w
        ws.freeze_panes = "A2"

    def _style_row(ws, er, alt):
        fill = ALT_FILL if alt else None
        for cell in ws[er]:
            cell.font      = ROW_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = BORDER
            if fill:
                cell.fill  = fill

    wb = Workbook()

    # ── Hoja 1: Registro de Amenazas ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Registro de Amenazas"
    # Usar siempre el orden de columnas que viene del frontend (leído del Excel).
    # Nunca imponer _THREATS_COLS para no destruir columnas del archivo existente.
    cols1 = list(threats[0].keys()) if threats else _THREATS_COLS
    ws1.append(cols1)
    _style_header(ws1, T_WIDTHS)
    for idx, rec in enumerate(threats):
        row = [_clean(rec.get(c)) for c in cols1]
        ws1.append(row)
        _style_row(ws1, idx + 2, idx % 2 == 1)

    # ── Hoja 2: Detalle de IoCs (doble encabezado) ───────────────────────────
    ws2 = wb.create_sheet("Detalle de IoCs")

    # Fila 1: encabezados fusionados
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "DETALLE DE INDICADORES DE COMPROMISO (IoC) — ISO 27001 A.5.7"
    ws2.merge_cells("D1:E1")
    ws2["D1"] = "ACCIONES DE BLOQUEO"
    for col_idx, fill in [(1, HDR_FILL), (4, HDR_FILL_GREEN)]:
        cell = ws2.cell(1, col_idx)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws2.row_dimensions[1].height = 28

    # Fila 2: nombres de columna
    for col_idx, val in enumerate(_IOCS_COLS, 1):
        cell = ws2.cell(2, col_idx)
        cell.value = val
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws2.row_dimensions[2].height = 22
    for i, w in enumerate(I_WIDTHS):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w
    ws2.freeze_panes = "A3"

    # Filas de datos desde fila 3
    for idx, rec in enumerate(iocs):
        row = [_clean(rec.get(c)) for c in _IOCS_COLS]
        ws2.append(row)
        _style_row(ws2, idx + 3, idx % 2 == 0)

    # ── Escritura atómica: temp → rename ─────────────────────────────────────
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=EXCEL_PATH.parent, suffix=".xlsx")
    import os; os.close(fd)
    try:
        wb.save(tmp)
        shutil.move(tmp, EXCEL_PATH)
    except Exception:
        try: os.unlink(tmp)
        except: pass
        raise


@app.route("/api/alerts", methods=["GET"])
@require_auth
def get_alerts():
    """Solo devuelve amenazas que Tenable confirmó como que afectan activos de la empresa."""
    try:
        if _USE_ORACLE:
            threats = _get_oracle().get_threats()
            afectan = [t for t in threats if str(t.get("¿Afecta Activos? (Tenable)", "")).startswith("SÍ")]
            return jsonify(afectan)

        if not EXCEL_PATH.exists():
            return jsonify([])
        import math
        df = pd.read_excel(EXCEL_PATH, sheet_name="Registro de Amenazas", engine="openpyxl")
        df = df.where(pd.notnull(df), None)
        afecta_col = next((c for c in df.columns if "afecta activos" in c.strip().lower()), None)
        if not afecta_col:
            afecta_col = next((c for c in df.columns if "comentarios soc" in c.strip().lower()), None)
        if not afecta_col:
            return jsonify([])
        if "afecta activos" in afecta_col.lower():
            afectan = df[df[afecta_col].astype(str).str.startswith("SÍ", na=False)]
        else:
            afectan = df[df[afecta_col].astype(str).str.contains("AFECTA LA EMPRESA", na=False)]
        records = afectan.to_dict(orient="records")
        return jsonify([
            {k: (None if isinstance(v, float) and math.isnan(v) else v) for k, v in row.items()}
            for row in records
        ])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/threats", methods=["GET"])
@require_auth
def get_threats():
    try:
        if _USE_ORACLE:
            rows = _get_oracle().get_threats()
            for row in rows:
                row.setdefault("CRITICIDAD", "")
                row.setdefault("PROBABILIDAD", "")
            return jsonify(rows)

        if not EXCEL_PATH.exists():
            return jsonify([])
        rows = _read_sheet("Registro de Amenazas")
        for row in rows:
            row.setdefault("CRITICIDAD", "")
            row.setdefault("PROBABILIDAD", "")
        return jsonify(rows)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/threats", methods=["PUT"])
@require_auth
def update_threats():
    try:
        body = request.get_json()
        if _USE_ORACLE:
            db = _get_oracle()
            db.save_threats_bulk(body)
            return jsonify({"ok": True})

        iocs = _read_sheet("Detalle de IoCs", header=_iocs_header()) if EXCEL_PATH.exists() else []
        _save_excel_atomic(body, iocs)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/iocs", methods=["GET"])
@require_auth
def get_iocs():
    try:
        if _USE_ORACLE:
            return jsonify(_get_oracle().get_iocs())

        if not EXCEL_PATH.exists():
            return jsonify([])
        return jsonify(_read_sheet("Detalle de IoCs", header=_iocs_header()))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/iocs", methods=["PUT"])
@require_auth
def update_iocs():
    try:
        body = request.get_json()
        if _USE_ORACLE:
            _get_oracle().save_iocs_bulk(body)
            return jsonify({"ok": True})

        threats = _read_sheet("Registro de Amenazas") if EXCEL_PATH.exists() else []
        _save_excel_atomic(threats, body)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/save", methods=["POST"])
@require_auth
def save_archive():
    """Endpoint principal: guarda amenazas + IoCs en una sola operación atómica."""
    try:
        body = request.get_json()
        if _USE_ORACLE:
            db = _get_oracle()
            db.save_threats_bulk(body.get("threats", []))
            db.save_iocs_bulk(body.get("iocs", []))
            return jsonify({"ok": True})

        _save_excel_atomic(body.get("threats", []), body.get("iocs", []))
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/migrate", methods=["POST"])
@require_auth
def migrate_excel():
    """Migra el Excel al formato de columnas actual (renombra columnas antiguas y reordena)."""
    if not EXCEL_PATH.exists():
        return jsonify({"error": "Sin archivo Excel"}), 404
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from src.excel_manager import _migrate_columns
        _migrate_columns()
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/download")
@require_auth
def download_excel():
    if not EXCEL_PATH.exists():
        return jsonify({"error": "Sin archivo Excel"}), 404
    return send_file(EXCEL_PATH, as_attachment=True, download_name="Informe_Amenazas.xlsx")


# ── Config ────────────────────────────────────────────────────────────────────
_CONFIG_KEYS = {
    "GROQ_API_KEY":       "CRITICAL",
    "GROQ_API_KEY_2":     "CRITICAL",
    "VT_API_KEY":         "HIGH",
    "TENABLE_ACCESS_KEY": "HIGH",
    "TENABLE_SECRET_KEY": "HIGH",
    "EMAIL_USER":         "MEDIUM",
    "SMTP_HOST":          "LOW",
    "REPORT_TO_EMAIL":    "LOW",
    "COMPANY_TECHS":      "LOW",
}


@app.route("/api/config", methods=["GET"])
@require_auth
def get_config():
    result = {}
    for key, severity in _CONFIG_KEYS.items():
        val = os.getenv(key, "")
        result[key] = {
            "severity": severity,
            "set":      bool(val),
            "masked":   val[:4] + "••••••••" if len(val) > 4 else "",
        }
    return jsonify({
        "keys":         result,
        "groq_model":   "llama-3.3-70b-versatile",
        "company_techs": os.getenv("COMPANY_TECHS", ""),
        "report_to":    os.getenv("REPORT_TO_EMAIL", ""),
    })


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/api/dashboard", methods=["GET"])
@require_auth
def get_dashboard():
    empty = {
        "threats": {"total": 0, "active": 0, "by_category": {}, "by_macro": {}, "by_status": {}, "by_month": {}, "company_impact": 0},
        "iocs":    {"total": 0, "by_type": {}},
    }
    if not EXCEL_PATH.exists():
        return jsonify(empty)
    try:
        import math
        threats_df = pd.read_excel(EXCEL_PATH, sheet_name="Registro de Amenazas", engine="openpyxl")
        iocs_df    = pd.read_excel(EXCEL_PATH, sheet_name="Detalle de IoCs",      engine="openpyxl", header=_iocs_header())

        threats_df = threats_df.where(pd.notnull(threats_df), None)
        iocs_df    = iocs_df.where(pd.notnull(iocs_df), None)

        # Normalize column names for lookup (strip + lower)
        threats_cols = {c.strip().lower(): c for c in threats_df.columns}
        iocs_cols    = {c.strip().lower(): c for c in iocs_df.columns}

        def _col(lookup: dict, *keys: str):
            for k in keys:
                if k.lower() in lookup:
                    return lookup[k.lower()]
            return None

        total  = len(threats_df)

        # "Activas" = Tenable confirmó que afectan activos (nueva col) o legado
        afecta_col = _col(threats_cols, "¿Afecta Activos? (Tenable)")
        soc_col    = _col(threats_cols, "Comentarios SOC")
        if afecta_col:
            active = int(threats_df[afecta_col].astype(str).str.startswith("SÍ", na=False).sum())
        elif soc_col:
            active = int(threats_df[soc_col].astype(str).str.contains("AFECTA LA EMPRESA", na=False).sum())
        else:
            active = 0

        def counts(col, *aliases):
            real = _col(threats_cols, col, *aliases)
            if not real:
                return {}
            return {k: v for k, v in threats_df[real].dropna().value_counts().items()
                    if not (isinstance(k, float) and math.isnan(k))}

        def month_counts():
            real = _col(threats_cols, "Fecha Detección", "Fecha Deteccion")
            if not real:
                return {}
            dates = pd.to_datetime(threats_df[real], errors="coerce").dropna()
            return dates.dt.to_period("M").astype(str).value_counts().sort_index().to_dict()

        # Sin gestionar = sin área responsable asignada (nueva) o Estado Activo (legado)
        area_col   = _col(threats_cols, "Área Responsable Remediación", "AREA RESPONSABLE DE REMEDIACION")
        estado_col = _col(threats_cols, "Estado")
        if area_col:
            company_impact = int(threats_df[area_col].isna().sum() + (threats_df[area_col].astype(str).str.strip() == '').sum())
        elif estado_col:
            company_impact = int((threats_df[estado_col].astype(str).str.strip() == "Activo").sum())
        else:
            company_impact = 0

        # IoC type column: nuevo "Tipo de IoC" o legado "TIPO DE IOC"
        ioc_type_col = _col(iocs_cols, "Tipo de IoC", "TIPO DE IOC", "tipo de ioc")
        ioc_by_type  = {k: v for k, v in iocs_df[ioc_type_col].dropna().value_counts().items()
                        if not (isinstance(k, float) and math.isnan(k))} if ioc_type_col else {}

        return jsonify({
            "threats": {
                "total":          total,
                "active":         active,
                "by_category":    counts("V-A", "Tipo de Amenaza"),
                "by_macro":       counts("NOMBRE", "Vulnerabilidad / Amenaza"),
                "by_status":      counts("¿Afecta Activos? (Tenable)", "Estado"),
                "by_month":       month_counts(),
                "company_impact": company_impact,
                "by_source":      counts("Fuente", "Fuente de Detección"),
            },
            "iocs": {
                "total":   len(iocs_df),
                "by_type": ioc_by_type,
            },
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


if __name__ == "__main__":
    debug = not _IS_PROD
    app.run(debug=debug, host="127.0.0.1", port=int(os.getenv("PORT", 5001)), threaded=True)
