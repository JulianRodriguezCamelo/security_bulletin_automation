import base64
import json
import os
import subprocess
import sys
from datetime import datetime
from functools import wraps
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, request, send_file, session
from flask_cors import CORS

load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))
from auth_manager import AuthManager

DIST = Path(__file__).parent / "frontend" / "dist"
app = Flask(__name__, static_folder=str(DIST), static_url_path="")
app.secret_key = os.getenv("SECRET_KEY", "argos-dev-secret-change-in-prod")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True

CORS(app, supports_credentials=True, origins=["http://localhost:5173", "http://127.0.0.1:5173"])

BULLETINS_DIR = Path.home() / "Documents" / "Casos_inteligencia_de_amenazas"
EXCEL_PATH    = BULLETINS_DIR / "Informe_Amenazas.xlsx"
MAIN_SCRIPT   = Path(__file__).parent / "main.py"

auth = AuthManager()


# ── Auth guard ────────────────────────────────────────────────────────────────
def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("authenticated"):
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return wrapper


# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
def login():
    data     = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not auth.has_users():
        auth.create_user(username, password)
        session["temp_username"] = username
        secret = auth.generate_totp_secret(username)
        qr     = auth.get_qr_image(username, secret)
        return jsonify({
            "step":   "totp_setup",
            "qr":     base64.b64encode(qr).decode(),
            "secret": secret,
        })

    if not auth.verify_password(username, password):
        return jsonify({"error": "Credenciales incorrectas"}), 401

    session["temp_username"] = username

    if not auth.has_totp(username):
        secret = auth.generate_totp_secret(username)
        qr     = auth.get_qr_image(username, secret)
        return jsonify({
            "step":   "totp_setup",
            "qr":     base64.b64encode(qr).decode(),
            "secret": secret,
        })

    return jsonify({"step": "totp"})


@app.route("/api/auth/totp/verify", methods=["POST"])
def verify_totp():
    data     = request.get_json()
    username = session.get("temp_username")
    code     = data.get("code", "").strip()

    if not username:
        return jsonify({"error": "No hay sesión pendiente"}), 400

    if auth.verify_totp(username, code):
        session.pop("temp_username", None)
        session["authenticated"] = True
        session["username"]      = username
        return jsonify({"step": "done", "username": username})

    return jsonify({"error": "Código incorrecto"}), 401


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/me")
def me():
    if session.get("authenticated"):
        return jsonify({"authenticated": True, "username": session["username"]})
    return jsonify({"authenticated": False})


# ── Files ─────────────────────────────────────────────────────────────────────
@app.route("/api/files", methods=["GET"])
@require_auth
def list_files():
    BULLETINS_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(BULLETINS_DIR.glob("*.pdf"))
    return jsonify([{"name": f.name, "size": f.stat().st_size} for f in files])


@app.route("/api/files", methods=["POST"])
@require_auth
def upload_files():
    BULLETINS_DIR.mkdir(parents=True, exist_ok=True)
    uploaded = []
    for file in request.files.getlist("files"):
        dest = BULLETINS_DIR / file.filename
        if dest.exists():
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = BULLETINS_DIR / f"{ts}_{file.filename}"
        file.save(dest)
        uploaded.append(dest.name)
    return jsonify({"uploaded": uploaded})


@app.route("/api/files/<name>", methods=["DELETE"])
@require_auth
def delete_file(name):
    path = BULLETINS_DIR / name
    if path.exists() and path.suffix.lower() == ".pdf":
        path.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404


# ── Process (SSE streaming) ───────────────────────────────────────────────────
# EventSource only supports GET, so we accept GET here.
@app.route("/api/process", methods=["GET"])
@require_auth
def process():
    def generate():
        try:
            proc = subprocess.Popen(
                [sys.executable, str(MAIN_SCRIPT)],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                cwd=str(MAIN_SCRIPT.parent),
            )
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    yield f"data: {json.dumps({'line': line})}\n\n"
            proc.wait()
            yield f"data: {json.dumps({'done': True, 'code': proc.returncode})}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Archive ───────────────────────────────────────────────────────────────────
def _read_sheet(sheet_name: str):
    import json as _json
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, engine="openpyxl")
    return _json.loads(df.to_json(orient="records", force_ascii=False, date_format="iso"))


def _write_sheet(sheet_name: str, records: list):
    from openpyxl import Workbook, load_workbook
    df = pd.DataFrame(records)
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append([None if isinstance(v, float) and pd.isna(v) else v for v in row])
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    wb.save(EXCEL_PATH)


@app.route("/api/archive/threats", methods=["GET"])
@require_auth
def get_threats():
    if not EXCEL_PATH.exists():
        return jsonify([])
    try:
        return jsonify(_read_sheet("Registro de Amenazas"))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/threats", methods=["PUT"])
@require_auth
def update_threats():
    try:
        _write_sheet("Registro de Amenazas", request.get_json())
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/iocs", methods=["GET"])
@require_auth
def get_iocs():
    if not EXCEL_PATH.exists():
        return jsonify([])
    try:
        return jsonify(_read_sheet("Detalle de IoCs"))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/iocs", methods=["PUT"])
@require_auth
def update_iocs():
    try:
        _write_sheet("Detalle de IoCs", request.get_json())
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/archive/download")
@require_auth
def download_excel():
    if not EXCEL_PATH.exists():
        return jsonify({"error": "Sin archivo Excel"}), 404
    return send_file(EXCEL_PATH, as_attachment=True, download_name="Informe_Amenazas.xlsx")


# ── Dashboard ─────────────────────────────────────────────────────────────────
import re as _re

_MESES = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12',
}

def _parse_es_month(s: str):
    m = _re.search(r'\d+\s+de\s+(\w+)\s+de\s+(\d{4})', str(s).lower())
    if m:
        month = _MESES.get(m.group(1))
        if month:
            return f"{m.group(2)}-{month}"
    return None

_MACRO_MAP = [
    ('Vulnerabilidad',  ['vulnerabilid']),
    ('Malware',         ['malware', 'ransomware', 'troyano', 'trojan', 'rat ', 'rat(', 'infostealer', 'backdoor', 'lofystealer', 'stealer', 'botn']),
    ('Phishing',        ['phishing']),
    ('APT',             ['apt']),
    ('Campaña',         ['campa', 'espionaje']),
    ('Poisoning',       ['poisoning', 'envenenamiento']),
    ('Compromiso',      ['compromiso']),
    ('Explotación',     ['explotaci', 'exploit']),
]

def _macro_type(val: str) -> str:
    v = val.lower()
    for label, keywords in _MACRO_MAP:
        if any(k in v for k in keywords):
            return label
    return 'Otro'


@app.route("/api/dashboard", methods=["GET"])
@require_auth
def dashboard():
    result = {
        "threats": {
            "total": 0, "active": 0,
            "by_category": {},   # Vulnerabilidad / Amenaza column
            "by_macro": {},      # Tipo de Amenaza grouped into macro categories
            "by_status": {},
            "by_month": {},
            "company_impact": 0,
        },
        "iocs": {"total": 0, "by_type": {}},
    }
    if not EXCEL_PATH.exists():
        return jsonify(result)

    try:
        df_t = pd.read_excel(EXCEL_PATH, sheet_name="Registro de Amenazas", engine="openpyxl")
        result["threats"]["total"] = len(df_t)

        # "Vulnerabilidad / Amenaza" → clean category donut
        cat_col = next((c for c in df_t.columns if "vulnerabilidad" in c.lower() and "amenaza" in c.lower()), None)
        if cat_col:
            result["threats"]["by_category"] = {
                str(k): int(v) for k, v in df_t[cat_col].value_counts().items()
            }

        # "Tipo de Amenaza" → grouped macro-categories bar chart
        tipo_col = next((c for c in df_t.columns if c.lower().startswith("tipo")), None)
        if tipo_col:
            macros: dict[str, int] = {}
            for val in df_t[tipo_col].dropna().astype(str):
                key = _macro_type(val)
                macros[key] = macros.get(key, 0) + 1
            result["threats"]["by_macro"] = dict(sorted(macros.items(), key=lambda x: -x[1]))

        # Estado
        status_col = next((c for c in df_t.columns if c.lower() == "estado"), None)
        if status_col:
            result["threats"]["by_status"] = {
                str(k): int(v) for k, v in df_t[status_col].value_counts().items()
            }
            result["threats"]["active"] = int(
                df_t[status_col].astype(str).str.lower().str.contains("activ", na=False).sum()
            )

        # Fecha en español → YYYY-MM
        date_col = next((c for c in df_t.columns if "fecha" in c.lower()), None)
        if date_col:
            months: dict[str, int] = {}
            for val in df_t[date_col].dropna().astype(str):
                ym = _parse_es_month(val)
                if ym:
                    months[ym] = months.get(ym, 0) + 1
            result["threats"]["by_month"] = dict(sorted(months.items()))

        # Impacto empresa
        impact_col = next((c for c in df_t.columns if "impacto" in c.lower() or "posible" in c.lower()), None)
        if impact_col:
            result["threats"]["company_impact"] = int(
                df_t[impact_col].astype(str).str.upper().str.strip().eq("SI").sum()
            )
    except Exception:
        pass

    try:
        df_i = pd.read_excel(EXCEL_PATH, sheet_name="Detalle de IoCs", engine="openpyxl")
        result["iocs"]["total"] = len(df_i)
        tipo_col = next((c for c in df_i.columns if "tipo" in c.lower()), None)
        if tipo_col:
            result["iocs"]["by_type"] = {
                str(k): int(v) for k, v in df_i[tipo_col].value_counts().items()
            }
    except Exception:
        pass

    return jsonify(result)


# ── Config ────────────────────────────────────────────────────────────────────
_CONFIG_KEYS = {
    "GROQ_API_KEY":       "CRITICAL",
    "VT_API_KEY":         "HIGH",
    "TENABLE_ACCESS_KEY": "HIGH",
    "TENABLE_SECRET_KEY": "HIGH",
    "EMAIL_USER":         "MEDIUM",
    "SMTP_HOST":          "LOW",
    "REPORT_TO_EMAIL":    "LOW",
    "COMPANY_TECHS":      "LOW",
}


@app.route("/api/config", methods=["GET"])
@require_auth
def get_config():
    result = {}
    for key, severity in _CONFIG_KEYS.items():
        val = os.getenv(key, "")
        result[key] = {
            "severity": severity,
            "set":      bool(val),
            "masked":   val[:4] + "••••••••" if len(val) > 4 else "",
        }
    return jsonify({
        "keys":         result,
        "groq_model":   "llama-3.3-70b-versatile",
        "company_techs": os.getenv("COMPANY_TECHS", ""),
        "report_to":    os.getenv("REPORT_TO_EMAIL", ""),
    })


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_spa(path: str):
    """Serve the React SPA for all non-API routes."""
    full = DIST / path
    if full.exists() and full.is_file():
        return app.send_static_file(path)
    return app.send_static_file("index.html")


if __name__ == "__main__":
    # use_reloader=False prevents the reloader from killing SSE streams mid-flight
    app.run(debug=True, port=5000, threaded=True, use_reloader=False)
