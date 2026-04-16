"""
reanalyze_vuln_field.py
-----------------------
Re-analyzes the "Vulnerabilidad / Amenaza" column in the existing Excel report.
For each row it calls the AI with the available context (Tipo de Amenaza +
Descripción + Descripción del Riesgo) and re-classifies the value as one of:
    Vulnerabilidad | Amenaza | Amenaza/Vulnerabilidad

The script updates the real report at:
    ~/Documents/Casos_inteligencia_de_amenazas/Informe_Amenazas.xlsx
"""

import os
import json
import time
import logging
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from groq import Groq

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

VALID_VALUES = {"Vulnerabilidad", "Amenaza", "Amenaza/Vulnerabilidad"}

REPORT_PATH = os.path.expanduser(
    r"~\Documents\Casos_inteligencia_de_amenazas\Informe_Amenazas.xlsx"
)


def classify_with_ai(client, model_id, threat_type: str, description: str, risk_description: str) -> str:
    """Calls AI to re-classify the Vulnerabilidad/Amenaza field."""
    prompt = f"""Eres un analista experto en ciberseguridad. Dado el siguiente contexto de un boletín de seguridad, clasifica el tipo exacto de la amenaza/vulnerabilidad.

Tipo de Amenaza: {threat_type}
Descripción: {description}
Descripción del Riesgo: {risk_description}

Responde ÚNICAMENTE con uno de estos tres valores exactos (sin comillas, sin texto extra):
- Vulnerabilidad   (si el boletín trata principalmente una falla técnica explotable en software/hardware)
- Amenaza          (si trata principalmente de un actor malicioso, malware o campaña activa sin vulnerabilidad específica)
- Amenaza/Vulnerabilidad  (si combina ambos: una amenaza activa que explota una vulnerabilidad concreta)

Tu respuesta debe ser exactamente una de esas tres opciones."""

    try:
        resp = client.chat.completions.create(
            messages=[{"role": "user", "content": prompt}],
            model=model_id,
            temperature=0.1,
        )
        result = resp.choices[0].message.content.strip().strip('"').strip("'")
        # Normalize and validate
        for v in VALID_VALUES:
            if v.lower() == result.lower():
                return v
        # If AI returned something unexpected, keep original
        logging.warning(f"Unexpected AI response '{result}', skipping.")
        return None
    except Exception as e:
        logging.error(f"AI call failed: {e}")
        return None


def main():
    load_dotenv()
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise ValueError("GROQ_API_KEY not found in .env")

    client = Groq(api_key=api_key)
    model_id = "llama-3.3-70b-versatile"

    if not os.path.exists(REPORT_PATH):
        logging.error(f"Excel not found at: {REPORT_PATH}")
        return

    logging.info(f"Loading: {REPORT_PATH}")
    wb = load_workbook(REPORT_PATH)

    # Support both possible sheet names
    sheet_name = None
    for candidate in ["Registro de Amenazas", "Informe de Amenazas"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break

    if not sheet_name:
        logging.error(f"No recognized sheet found. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # Find column indexes by header name
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    logging.info(f"Columns found: {list(headers.keys())}")

    col_vuln = headers.get("Vulnerabilidad / Amenaza")
    col_type = headers.get("Tipo de Amenaza")
    col_desc = headers.get("Descripción")
    col_risk = headers.get("Descripción del Riesgo")

    if not col_vuln:
        logging.error("Column 'Vulnerabilidad / Amenaza' not found.")
        return

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(row_idx, 1).value
        if not row_id:
            continue

        threat_type   = ws.cell(row_idx, col_type).value or "" if col_type else ""
        description   = ws.cell(row_idx, col_desc).value or "" if col_desc else ""
        risk_desc     = ws.cell(row_idx, col_risk).value or "" if col_risk else ""
        current_value = ws.cell(row_idx, col_vuln).value or ""

        logging.info(f"[{row_id}] Tipo: {threat_type[:60]} | Actual: {current_value}")

        new_value = classify_with_ai(client, model_id, threat_type, description, risk_desc)

        if new_value and new_value != current_value:
            ws.cell(row_idx, col_vuln).value = new_value
            logging.info(f"  -> Updated: '{current_value}' => '{new_value}'")
            updated += 1
        elif new_value:
            logging.info(f"  -> No change needed ({new_value})")
        else:
            logging.warning(f"  -> Could not classify, keeping '{current_value}'")

        time.sleep(2)  # Avoid Groq rate limit

    # Re-apply data validation with the correct allowed values
    dv = DataValidation(
        type="list",
        formula1='"Vulnerabilidad,Amenaza,Amenaza/Vulnerabilidad"',
        allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col_vuln)}2:{get_column_letter(col_vuln)}{max(100, ws.max_row)}")

    wb.save(REPORT_PATH)
    logging.info(f"Done. {updated} row(s) updated. File saved: {REPORT_PATH}")


if __name__ == "__main__":
    main()
