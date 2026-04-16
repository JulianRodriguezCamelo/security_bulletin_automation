import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Column definitions ──────────────────────────────────────────────────────
AMENAZAS_COLUMNS = [
    ("ID",                   15),
    ("Fecha Detección",      22),
    ("Tipo de Amenaza",      40),
    ("Fuente de Detección",  25),
    ("Vulnerabilidad / Amenaza", 25),
    ("Descripción",          55),
    ("Descripción del Riesgo", 45),
    ("Posible Impacto",      18),
    ("Indicadores (IoC)",    35),
    ("TTP (Técnicas, Tácticas y Procedimientos)", 45),
    ("Acción Recomendada",   55),
    ("Reportó",              20),
    ("Comentarios SOC",      35),
    ("Estado",               18),
    ("Comentarios FIDU",     35),
    ("BLOQUEO ANTIVIRUS",    22),
    ("BLOQUEO FIREWALL",     22),
    ("CASO FIREWALL",        22),
]

IOCS_COLUMNS = [
    ("TIPO DE IOC", 18),
    ("IOC",         35),
    ("FECHA",       18),
    ("ANTIVIRUS",   12),
    ("FIREWALL",    12),
    ("CASO FIREWALL", 22),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_ROW_FILL  = PatternFill("solid", fgColor="D9E1F2")   # soft blue
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")

def _fmt_fecha(dt: datetime) -> str:
    """Returns date formatted in Spanish long style."""
    dias   = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]
    meses  = ["enero","febrero","marzo","abril","mayo","junio",
               "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    return f"{dias[dt.weekday()]}, {dt.day} de {meses[dt.month-1]} de {dt.year}"

def _make_id(row_number: int) -> str:
    return f"WX-{row_number:02d}"

class ExcelManager:
    """
    Generates a security report with two sheets: 
    - Registro de Amenazas (18 columns)
    - Detalle de IoCs (6 columns)
    """

    def __init__(self, output_path: str = "./data/Informe_Amenazas.xlsx"):
        self.output_path = output_path
        self.records: list[dict] = []

    def add_record(self, threat_data: dict):
        """Append an analysis result dict to the report list."""
        self.records.append(threat_data)

    def save(self):
        """Creates or appends to a multi-sheet bitácora."""
        if not self.records:
            return

        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        
        if os.path.exists(self.output_path):
            try:
                wb = load_workbook(self.output_path)
            except:
                wb = Workbook()
        else:
            wb = Workbook()

        # ── Part 1: Registro de Amenazas ───────────────────────────────
        sheet_name_am = "Registro de Amenazas"
        if sheet_name_am in wb.sheetnames:
            ws_am = wb[sheet_name_am]
        else:
            # If default sheet is 'Sheet', rename it, else create new
            if wb.active.title == "Sheet":
                ws_am = wb.active
                ws_am.title = sheet_name_am
            else:
                ws_am = wb.create_sheet(sheet_name_am)
            self._write_headers(ws_am, AMENAZAS_COLUMNS)

        start_row_am = ws_am.max_row + 1
        last_id_num = 0
        existing_ids = set()
        
        for r in range(2, start_row_am):
            cell_val = ws_am.cell(r, 1).value
            if cell_val:
                val_str = str(cell_val).strip().upper()
                existing_ids.add(val_str)
                if "-" in val_str:
                    try:
                        num = int(val_str.split("-")[-1])
                        if num > last_id_num:
                            last_id_num = num
                    except: pass

        # ── Part 2: Detalle de IoCs ─────────────────────────────────────
        sheet_name_ioc = "Detalle de IoCs"
        if sheet_name_ioc in wb.sheetnames:
            ws_ioc = wb[sheet_name_ioc]
        else:
            ws_ioc = wb.create_sheet(sheet_name_ioc)
            self._write_headers(ws_ioc, IOCS_COLUMNS)
        
        start_row_ioc = ws_ioc.max_row + 1

        # ── Write Data Loop ─────────────────────────────────────────────
        current_date_str = datetime.now().strftime("%d/%m/%Y")
        
        added_count = 0
        for idx, record in enumerate(self.records):
            raw_id = record.get("id")
            if raw_id and not str(raw_id).upper().startswith("WX-"):
                display_id = f"WX-{raw_id}".upper()
            else:
                display_id = (raw_id or _make_id(last_id_num + added_count + 1)).upper()

            if display_id in existing_ids:
                # Evitar duplicados
                continue
            
            existing_ids.add(display_id)
            
            # A. Threat Record
            row_idx = start_row_am + added_count
            fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
            
            added_count += 1
            
            # Build VT-enriched IoC list for the main sheet
            vt_results = record.get("vt_results", {})
            ioc_list_text = []
            for ioc in record.get("iocs", []):
                # Clean before display lookup if needed
                vt = vt_results.get(ioc) or vt_results.get(ioc.replace("[.]", "."))
                if vt and vt.get("malicious", 0) > 0:
                    ioc_list_text.append(f"{ioc} (VT: {vt['malicious']} malicioso)")
                else:
                    ioc_list_text.append(ioc)

            row_values_am = [
                display_id,
                _fmt_fecha(datetime.now()),
                record.get("threat_type") or "",
                record.get("source") or "",
                record.get("vulnerability_name") or "",
                record.get("description") or "",
                record.get("risk_description") or "",
                record.get("possible_impact") or "",
                "\n".join(ioc_list_text) if ioc_list_text else "N/A",
                "\n".join(record.get("ttps", [])) if record.get("ttps") else "N/A",
                record.get("recommended_action") or "",
                "WEXLER",
                record.get("soc_comments") or "",
                record.get("status", "Gestionado"),
                record.get("fidu_comments") or "",
                record.get("antivirus_block", "NO APLICA"),
                record.get("firewall_block", "      "),
                record.get("firewall_case", "     "),
            ]
            self._write_row(ws_am, row_idx, row_values_am, fill)

            # B. IoC Inventory Rows
            for ioc_name, vt_data in vt_results.items():
                ioc_row_idx = ws_ioc.max_row + 1
                ioc_fill = ALT_ROW_FILL if ioc_row_idx % 2 == 0 else None
                
                # Use checkmark emoji as requested by user
                av_check = "✅" if vt_data.get("malicious", 0) > 0 else "N/A"
                fw_check = "✅" if vt_data.get("malicious", 0) > 5 else "N/A" # Example logic

                row_values_ioc = [
                    vt_data.get("type", "Hash/Otro"),
                    ioc_name,
                    current_date_str,
                    av_check,
                    fw_check,
                    "" # Firewall case empty
                ]
                self._write_row(ws_ioc, ioc_row_idx, row_values_ioc, ioc_fill)

        # ── Styling fixes ──────────────────────────────────────────────
        for sheet in [ws_am, ws_ioc]:
            sheet.freeze_panes = "A2"
            last_col = get_column_letter(sheet.max_column)
            sheet.auto_filter.ref = f"A1:{last_col}1"

        # Validación de datos para la columna "Vulnerabilidad / Amenaza" (Columna E)
        dv = DataValidation(type="list", formula1='"Vulnerabilidad,Amenaza,Amenaza/Vulnerabilidad"', allow_blank=True)
        ws_am.add_data_validation(dv)
        dv.add(f"E2:E{max(100, ws_am.max_row)}")

        wb.save(self.output_path)

    def _write_headers(self, ws, columns):
        ws.row_dimensions[1].height = 38
        for col_idx, (header, width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_row(self, ws, row_idx, values, fill=None):
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if fill: cell.fill = fill
        ws.row_dimensions[row_idx].height = 45 if ws.title == "Detalle de IoCs" else 60
