from pathlib import Path

import pandas as pd
import streamlit as st

EXCEL_PATH    = Path.home() / "Documents" / "Casos_inteligencia_de_amenazas" / "Informe_Amenazas.xlsx"
SHEET_THREATS = "Registro de Amenazas"
SHEET_IOCS    = "Detalle de IoCs"

_TYPE_COLORS: dict[str, tuple[str, str]] = {
    "vulnerabilidad": ("#7c3aed", "rgba(124,58,237,0.10)"),
    "amenaza":        ("#C1294A", "rgba(193,41,74,0.10)"),
    "malware":        ("#DC2626", "rgba(220,38,38,0.10)"),
    "phishing":       ("#DB2777", "rgba(219,39,119,0.10)"),
    "ransomware":     ("#DC2626", "rgba(220,38,38,0.10)"),
    "apt":            ("#7c3aed", "rgba(124,58,237,0.10)"),
}

_STATUS_COLORS: dict[str, tuple[str, str]] = {
    "activ":    ("#C1294A", "rgba(193,41,74,0.08)"),
    "cerrad":   ("#16a34a", "rgba(22,163,74,0.08)"),
    "pendient": ("#D97706", "rgba(217,119,6,0.08)"),
    "en curso": ("#2563EB", "rgba(37,99,235,0.08)"),
}


def _load(sheet: str) -> pd.DataFrame | None:
    try:
        return pd.read_excel(EXCEL_PATH, sheet_name=sheet, engine="openpyxl")
    except Exception as exc:
        st.error(f"Error leyendo «{sheet}»: {exc}")
        return None


def _find_col(df: pd.DataFrame, kw: str) -> str | None:
    kw = kw.lower()
    for c in df.columns:
        if kw in c.lower():
            return c
    return None


def _save_excel(df_threats: pd.DataFrame | None, df_iocs: pd.DataFrame | None):
    import io
    from openpyxl import load_workbook

    buf = io.BytesIO()
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        from openpyxl import Workbook
        wb = Workbook()

    def _write_sheet(wb, sheet_name: str, df: pd.DataFrame):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))

    if df_threats is not None:
        _write_sheet(wb, SHEET_THREATS, df_threats)
    if df_iocs is not None:
        _write_sheet(wb, SHEET_IOCS, df_iocs)

    # remove default empty sheet if present
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_archive():
    st.markdown("""
    <style>
    /* ── Responsive archive ── */
    @media (max-width: 768px) {
        .kpi-grid { grid-template-columns: repeat(2,1fr) !important; }
        .filter-row { flex-direction: column !important; }
    }
    @media (max-width: 480px) {
        .kpi-grid { grid-template-columns: 1fr !important; }
    }
    /* Download button override – never black */
    [data-testid="stDownloadButton"] > button {
        background: #FFFFFF !important;
        color: #7A3045 !important;
        border: 1.5px solid #E0CDD3 !important;
        box-shadow: 0 1px 4px rgba(74,14,36,0.07) !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        padding: 10px 18px !important;
        transition: all 200ms !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        color: #C1294A !important;
        border-color: #C1294A !important;
        background: #FEF5F7 !important;
        box-shadow: 0 2px 10px rgba(193,41,74,0.14) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        """
        <div class="fade-up" style="margin-bottom:24px;">
            <h1 style="margin:0;font-size:1.4rem;font-weight:800;color:#1C0A0F;
                display:flex;align-items:center;gap:10px;">
                <span>📊</span> Archivo de Boletines
            </h1>
            <p style="font-size:13px;color:#A07080;margin:4px 0 0;">
                Registro histórico de amenazas e indicadores de compromiso (IoCs).
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not EXCEL_PATH.exists():
        st.markdown(
            """
            <div style="background:#FFFFFF;border:2px dashed #EAD8DE;
                border-radius:18px;padding:56px 32px;text-align:center;
                box-shadow:0 2px 12px rgba(74,14,36,0.06);">
                <div style="font-size:40px;margin-bottom:14px;opacity:0.25;">📁</div>
                <p style="font-size:14px;color:#7A4A55;margin:0;font-weight:600;">Sin datos todavía</p>
                <p style="font-size:12px;color:#C4A8B2;margin:8px 0 0;">
                    Procesa al menos un PDF desde
                    <strong style="color:#C1294A;">Procesar</strong>
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    # ── Load data into session state so edits persist ─────────────────────────
    if "archive_threats" not in st.session_state or st.button(
        "↺  Recargar desde archivo", key="reload_excel", type="secondary"
    ):
        df_t = _load(SHEET_THREATS)
        df_i = _load(SHEET_IOCS)
        st.session_state.archive_threats = df_t
        st.session_state.archive_iocs    = df_i

    df_threats = st.session_state.archive_threats
    df_iocs    = st.session_state.archive_iocs

    tab1, tab2 = st.tabs(["🗂️  Registro de Amenazas", "🔍  Detalle de IoCs"])

    with tab1:
        if df_threats is not None and not df_threats.empty:
            edited = _render_threats(df_threats)
            if edited is not None:
                st.session_state.archive_threats = edited
        else:
            st.info("Aún no hay registros.")

    with tab2:
        if df_iocs is not None and not df_iocs.empty:
            edited_ioc = _render_iocs(df_iocs)
            if edited_ioc is not None:
                st.session_state.archive_iocs = edited_ioc
        else:
            st.info("Aún no hay IoCs registrados.")

    # ── Footer: save + download ───────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)

    col_save, col_dl, col_info = st.columns([1, 1, 3])

    with col_save:
        if st.button("💾  Guardar cambios", key="save_excel", type="primary", use_container_width=True):
            try:
                raw = _save_excel(
                    st.session_state.get("archive_threats"),
                    st.session_state.get("archive_iocs"),
                )
                EXCEL_PATH.write_bytes(raw)
                st.success("✓ Cambios guardados en el archivo Excel")
            except Exception as exc:
                st.error(f"Error al guardar: {exc}")

    with col_dl:
        try:
            raw = _save_excel(
                st.session_state.get("archive_threats"),
                st.session_state.get("archive_iocs"),
            )
            st.download_button(
                "⬇  Descargar Excel",
                data=raw,
                file_name=EXCEL_PATH.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel",
            )
        except Exception as exc:
            st.error(f"Error preparando descarga: {exc}")

    with col_info:
        st.markdown(
            f'<p style="font-size:11px;color:#C4A8B2;padding-top:12px;word-break:break-all;">'
            f'{EXCEL_PATH}</p>',
            unsafe_allow_html=True,
        )


def _render_threats(df: pd.DataFrame) -> pd.DataFrame | None:
    status_col = _find_col(df, "estado")
    type_col   = _find_col(df, "tipo")
    date_col   = _find_col(df, "fecha")

    active_count = 0
    if status_col:
        active_count = int(df[status_col].astype(str).str.lower().str.contains("activ", na=False).sum())

    type_count = int(df[type_col].nunique()) if type_col else 0

    last_date = "—"
    if date_col:
        try:
            dt = pd.to_datetime(df[date_col], errors="coerce").max()
            if pd.notna(dt):
                last_date = dt.strftime("%d/%m/%Y")
        except Exception:
            pass

    kpi_items = [
        ("📋", "Total Amenazas",  str(len(df)),      "#C1294A"),
        ("⚡", "Activas",          str(active_count), "#DC2626"),
        ("🏷️", "Tipos distintos",  str(type_count),   "#7c3aed"),
        ("📅", "Último registro", last_date,         "#16a34a"),
    ]

    st.markdown(
        """<div class="kpi-grid" style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px;">"""
        + "".join(
            f"""
            <div style="background:#FFFFFF;border:1px solid #EAD8DE;border-top:3px solid {c};
                border-radius:14px;padding:16px;
                box-shadow:0 2px 10px rgba(74,14,36,0.07);" class="fade-up">
                <div style="font-size:18px;margin-bottom:6px;">{ic}</div>
                <div style="font-size:10px;color:#8A5060;font-weight:600;text-transform:uppercase;
                    letter-spacing:0.07em;">{lbl}</div>
                <div style="font-size:24px;font-weight:800;color:#1C0A0F;margin-top:4px;">{val}</div>
            </div>"""
            for ic, lbl, val, c in kpi_items
        )
        + "</div>",
        unsafe_allow_html=True,
    )

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.container():
        fc1, fc2, fc3 = st.columns([1, 1, 1])
        filtered = df.copy()

        with fc1:
            if type_col:
                opts = ["Todos"] + sorted(df[type_col].dropna().astype(str).unique().tolist())
                sel = st.selectbox("Tipo de amenaza", opts, key="th_type")
                if sel != "Todos":
                    filtered = filtered[filtered[type_col].astype(str) == sel]

        with fc2:
            if status_col:
                opts = ["Todos"] + sorted(df[status_col].dropna().astype(str).unique().tolist())
                sel = st.selectbox("Estado", opts, key="th_status")
                if sel != "Todos":
                    filtered = filtered[filtered[status_col].astype(str) == sel]

        with fc3:
            kw = st.text_input("Buscar", placeholder="ransomware, CVE-2024…", key="th_search")
            if kw:
                mask = pd.Series(False, index=filtered.index)
                for col in filtered.select_dtypes(include="object").columns:
                    mask |= filtered[col].astype(str).str.contains(kw, case=False, na=False)
                filtered = filtered[mask]

    st.markdown(
        f'<p style="font-size:11px;color:#B09AA5;margin:8px 0 4px;">'
        f'{len(filtered)} de {len(df)} registros — '
        f'<em>Edita directamente en la tabla y luego guarda</em></p>',
        unsafe_allow_html=True,
    )

    edited = st.data_editor(
        filtered,
        use_container_width=True,
        height=480,
        hide_index=True,
        num_rows="dynamic",
        column_config=_threat_column_config(filtered),
        key="editor_threats",
    )

    # Merge edits back into full df (matching by index)
    if edited is not None:
        df.update(edited)
        return df
    return None


def _threat_column_config(df: pd.DataFrame) -> dict:
    cfg = {}
    type_col   = _find_col(df, "tipo")
    status_col = _find_col(df, "estado")
    date_col   = _find_col(df, "fecha")
    if type_col:
        cfg[type_col]   = st.column_config.TextColumn(type_col, width="small")
    if status_col:
        cfg[status_col] = st.column_config.TextColumn(status_col, width="small")
    if date_col:
        cfg[date_col]   = st.column_config.DateColumn(date_col, format="DD/MM/YYYY", width="small")
    return cfg


def _render_iocs(df: pd.DataFrame) -> pd.DataFrame | None:
    type_col = _find_col(df, "tipo")
    av_col   = _find_col(df, "av")

    mal_count = 0
    if av_col:
        try:
            mal_count = int(df[av_col].astype(str).str.match(r"^\d+/\d+$").sum())
        except Exception:
            pass

    kpi_items = [
        ("🔍", "Total IoCs",    str(len(df)),                                           "#C1294A"),
        ("🏷️", "Tipos",         str(df[type_col].nunique()) if type_col else "—",       "#7c3aed"),
        ("🦠", "Con detec. AV", str(mal_count),                                         "#DC2626"),
    ]

    st.markdown(
        """<div class="kpi-grid" style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:20px;">"""
        + "".join(
            f"""
            <div style="background:#FFFFFF;border:1px solid #EAD8DE;border-top:3px solid {c};
                border-radius:14px;padding:16px;
                box-shadow:0 2px 10px rgba(74,14,36,0.07);" class="fade-up">
                <div style="font-size:18px;margin-bottom:6px;">{ic}</div>
                <div style="font-size:10px;color:#8A5060;font-weight:600;text-transform:uppercase;
                    letter-spacing:0.07em;">{lbl}</div>
                <div style="font-size:24px;font-weight:800;color:#1C0A0F;margin-top:4px;">{val}</div>
            </div>"""
            for ic, lbl, val, c in kpi_items
        )
        + "</div>",
        unsafe_allow_html=True,
    )

    fi1, fi2 = st.columns(2)
    filtered = df.copy()

    with fi1:
        ioc_col = _find_col(df, "ioc") or (df.columns[1] if len(df.columns) > 1 else df.columns[0])
        search  = st.text_input("Buscar IoC", placeholder="IP, dominio, hash…", key="ioc_search")
        if search:
            filtered = filtered[
                filtered[ioc_col].astype(str).str.contains(search, case=False, na=False)
            ]

    with fi2:
        if type_col:
            opts = ["Todos"] + sorted(df[type_col].dropna().astype(str).unique().tolist())
            sel  = st.selectbox("Tipo de IoC", opts, key="ioc_type")
            if sel != "Todos":
                filtered = filtered[filtered[type_col].astype(str) == sel]

    st.markdown(
        f'<p style="font-size:11px;color:#B09AA5;margin:8px 0 4px;">'
        f'{len(filtered)} de {len(df)} IoCs — '
        f'<em>Edita directamente en la tabla y luego guarda</em></p>',
        unsafe_allow_html=True,
    )

    edited = st.data_editor(
        filtered,
        use_container_width=True,
        height=420,
        hide_index=True,
        num_rows="dynamic",
        key="editor_iocs",
    )

    if edited is not None:
        df.update(edited)
        return df
    return None
